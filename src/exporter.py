from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import CompanyRecord


BASE_COLUMNS = [
    ("Nombre de empresa", "nombre_empresa"),
    ("Correo", "correo"),
    ("Teléfono 1", "telefono_1"),
    ("Teléfono 2", "telefono_2"),
    ("Dirección", "direccion"),
    ("Estado", "estado"),
    ("País", "pais"),
    ("Sitio web", "sitio_web"),
    ("LinkedIn", "linkedin"),
    ("Segmento", "segmento"),
]

CONTROL_COLUMNS = [
    ("Ciudad", "ciudad"),
    ("URL fuente", "url_fuente"),
    ("Fuente correo", "fuente_correo"),
    ("Fuente teléfono", "fuente_telefono"),
    ("Fuente sitio web", "fuente_sitio_web"),
    ("Fuente LinkedIn", "fuente_linkedin"),
    ("Estado extracción", "estado_extraccion"),
    ("Observaciones", "observaciones"),
]


def records_to_excel(records: Iterable[CompanyRecord], include_control: bool = False) -> bytes:
    records = list(records)
    columns = BASE_COLUMNS + (CONTROL_COLUMNS if include_control else [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(records, start=2):
        data = record.as_dict()
        for col_idx, (_, key) in enumerate(columns, start=1):
            value = data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in {"sitio_web", "linkedin", "url_fuente"} and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    if records:
        end_col = get_column_letter(len(columns))
        table = Table(displayName="TablaEmpresas", ref=f"A1:{end_col}{len(records)+1}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    widths = {
        "nombre_empresa": 36, "correo": 32, "telefono_1": 20, "telefono_2": 20,
        "direccion": 52, "estado": 24, "pais": 18, "sitio_web": 38,
        "linkedin": 45, "segmento": 28, "ciudad": 24, "url_fuente": 50,
        "fuente_correo": 18, "fuente_telefono": 18, "fuente_sitio_web": 18,
        "fuente_linkedin": 18, "estado_extraccion": 18, "observaciones": 48,
    }
    for idx, (_, key) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(key, 22)

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
